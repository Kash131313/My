#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import time
import sqlite3
import asyncio
import logging
import concurrent.futures
import tempfile
import shutil
from functools import partial
from datetime import datetime
from typing import List, Tuple, Optional

import pytz
from dateutil import tz
from telethon import TelegramClient, events, errors
from openpyxl import Workbook, load_workbook
from urllib.parse import urlsplit, urlunsplit

# ----------------------------- CONFIG -----------------------------
API_ID = int(os.environ.get('TG_API_ID', '34599696'))
API_HASH = os.environ.get('TG_API_HASH', '1499462b8ad56f15bf407582b7a2175a')
PHONE = os.environ.get('TG_PHONE', '+79256991300')
TWOFA_PASSWORD = os.environ.get('TG_2FA_PASSWORD', '')  # optional
TARGET_CHANNEL = os.environ.get('TG_TARGET_CHANNEL', '-1003432562272')

EXCEL_PATH = "C:/Users/Иван/Desktop/links.xlsx"
STATE_DB = 'state_links.db'
SESSION_NAME = 'telegram_link_parser.session'

# Источники
SOURCES = [
    -1003497455949,
]

# Интервалы/настройки
MAX_MESSAGES_POLL = 200
TIMEZONE = 'Europe/Moscow'
CHANNEL_FETCH_LIMIT = 5000  # сколько сообщений вытянуть из целевого канала при старте
DB_BUSY_RETRIES = 8

# Реакции, которые мы синхронизируем
REACTIONS = ['👍', '🔥', '🤝']

# Очистка канала при старте (если True — мы попытаемся очистить канал при старте; если False, мы попытаемся синхронизировать)
CLEAR_TARGET_CHANNEL_ON_START = True

# Логирование
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger('linkparser')

# Регекс для URL - базовый
URL_RE = re.compile(r"(https?://\S+)")
HYPERLINK_RE = re.compile(r'HYPERLINK\("([^"]+)"\s*,\s*"([^"]*)"\)', re.IGNORECASE)

# -------------------------- DB WORKER ------------------------------
DB_PATH = STATE_DB
_db_queue: asyncio.Queue = asyncio.Queue()
_executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
_db_conn_for_worker: Optional[sqlite3.Connection] = None

def _open_sqlite_conn_for_worker(path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(path, timeout=30, check_same_thread=False)
    try:
        conn.execute('PRAGMA journal_mode=WAL;')
    except Exception:
        pass
    try:
        conn.execute('PRAGMA synchronous=NORMAL;')
    except Exception:
        pass
    try:
        conn.execute('PRAGMA busy_timeout=5000;')
    except Exception:
        pass
    return conn

def _sqlite_exec(conn: sqlite3.Connection, sql: str, params: Tuple = ()):
    cur = conn.cursor()
    cur.execute(sql, params or ())
    if sql.strip().upper().startswith('INSERT'):
        conn.commit()
        return cur.lastrowid
    else:
        rows = cur.fetchall()
        conn.commit()
        return rows

async def db_worker():
    global _db_conn_for_worker
    loop = asyncio.get_running_loop()
    _db_conn_for_worker = _open_sqlite_conn_for_worker(DB_PATH)
    logger.info("DB worker started")
    while True:
        item = await _db_queue.get()
        if item is None:
            _db_queue.task_done()
            break
        sql, params, fut = item
        try:
            func = partial(_sqlite_exec, _db_conn_for_worker, sql, params)
            res = await loop.run_in_executor(_executor, func)
            if fut and not fut.cancelled():
                fut.set_result(res)
        except Exception as e:
            logger.exception("DB worker error executing SQL")
            if fut and not fut.cancelled():
                fut.set_exception(e)
        finally:
            _db_queue.task_done()
    try:
        _db_conn_for_worker.close()
    except Exception:
        pass
    logger.info("DB worker stopped")

async def enqueue_db(sql: str, params: Tuple = (), wait_result: bool = True):
    if wait_result:
        fut = asyncio.get_running_loop().create_future()
    else:
        fut = None
    await _db_queue.put((sql, params, fut))
    if fut:
        return await fut
    return None

async def db_fetchall(sql: str, params: Tuple = ()):
    res = await enqueue_db(sql, params, wait_result=True)
    return res or []

async def db_execute(sql: str, params: Tuple = ()):
    return await enqueue_db(sql, params, wait_result=True)

# ------------------------- EXCEL HELPERS -------------------------

def ensure_excel(path: str):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = 'links'
        ws['A1'] = 'link'
        ws['B1'] = 'time'
        wb.save(path)
        logger.info(f'Created new Excel file at {path}')

def format_time_for_excel(ts: int) -> str:
    dt = datetime.fromtimestamp(ts, tz=pytz.utc).astimezone(tz.gettz(TIMEZONE))
    return dt.strftime('%H:%M-%d.%m.%y')

def save_workbook_atomic(wb, target_path):
    dir_name = os.path.dirname(target_path) or "."
    fd, tmp_path = tempfile.mkstemp(dir=dir_name, prefix=".tmp_excel_", suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp_path)
        shutil.move(tmp_path, target_path)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass

async def rebuild_excel_from_db(path: str):
    loop = asyncio.get_running_loop()
    now = int(time.time())
    cutoff = now - 24 * 3600
    rows = await db_fetchall('SELECT url, first_seen FROM urls WHERE first_seen >= ? ORDER BY first_seen ASC', (cutoff,))
    try:
        def _write(rows_local):
            ensure_excel(path)
            wb = load_workbook(path)
            ws = wb.active
            if ws.max_row >= 2:
                ws.delete_rows(2, ws.max_row)
            r = 2
            for url, first_seen in rows_local:
                ws.cell(row=r, column=1).value = f'=HYPERLINK("{url}","{url}")'
                ws.cell(row=r, column=2).value = format_time_for_excel(first_seen)
                r += 1
            save_workbook_atomic(wb, path)
            return len(rows_local)
        cnt = await loop.run_in_executor(_executor, partial(_write, rows))
        logger.info(f'Excel rebuilt with {cnt} rows')
    except Exception as e:
        logger.exception(f'Cannot rebuild Excel: {e}')

# ------------------------- URL / NORM HELPERS ---------------------

def normalize_url(u: str) -> str:
    if not u:
        return u
    u = u.strip().rstrip('.,;:!?\'"')
    has_scheme = '://' in u
    parse_target = u if has_scheme else ('http://' + u)
    try:
        parts = urlsplit(parse_target)
    except Exception:
        return u
    scheme = parts.scheme.lower() if parts.scheme else 'http'
    netloc = parts.netloc.lower()
    if netloc.startswith('www.'):
        netloc = netloc[4:]
    if ':' in netloc:
        host, port = netloc.split(':', 1)
        if port in ('80','443'):
            netloc = host
    path = parts.path or '/'
    query = parts.query or ''
    frag = ''
    return urlunsplit((scheme, netloc, path, query, frag))

# -------------------------- TELETHON APP -------------------------

client = TelegramClient(SESSION_NAME, API_ID, API_HASH)

async def send_channel_message(url: str, msg_ts: int) -> Optional[int]:
    try:
        try:
            target = int(TARGET_CHANNEL)
        except Exception:
            target = TARGET_CHANNEL
        text = f'{url}\n{format_time_for_excel(msg_ts)}'
        msg = await client.send_message(entity=target, message=text)
        logger.info(f'Sent to channel {target} msg_id={msg.id} url={url}')
        return msg.id
    except Exception as e:
        logger.exception(f'Failed to send to channel: {e}')
        return None

async def delete_channel_message(msg_id: int):
    if not msg_id:
        return
    try:
        try:
            target = int(TARGET_CHANNEL)
        except Exception:
            target = TARGET_CHANNEL
        await client.delete_messages(entity=target, message_ids=[msg_id])
        logger.info(f'Deleted message in channel id={msg_id}')
    except Exception as e:
        logger.warning(f'Failed to delete channel message {msg_id}: {e}')

async def try_send_reaction(peer, msg_id, reaction: str) -> bool:
    try:
        if hasattr(client, 'send_reaction'):
            await client.send_reaction(entity=peer, message=msg_id, reaction=reaction)
            return True
        return False
    except errors.RPCError as e:
        logger.warning(f'RPC error sending reaction {reaction} to {peer}:{msg_id} -> {e}')
        return False
    except Exception as e:
        logger.warning(f'Cannot send reaction {reaction} to {peer}:{msg_id} -> {e}')
        return False

# ---------------------- Reaction sync / helpers --------------------

async def gather_channel_messages_with_urls(limit=CHANNEL_FETCH_LIMIT):
    try:
        try:
            target = int(TARGET_CHANNEL)
        except Exception:
            target = TARGET_CHANNEL
        msgs = await client.get_messages(entity=target, limit=limit)
    except Exception as e:
        logger.exception(f'Failed to fetch channel messages: {e}')
        return []
    out = []
    for m in msgs:
        if not m or not getattr(m, 'message', None):
            continue
        text = m.message
        found = URL_RE.findall(text)
        if not found:
            continue
        url_text = found[0].strip()
        norm = normalize_url(url_text)
        try:
            ts = int(m.date.replace(tzinfo=pytz.utc).timestamp())
        except Exception:
            ts = 0
        out.append((m.id, text, norm, ts, getattr(m, 'reactions', None)))
    out = list(reversed(out))  # oldest -> newest
    return out

def parse_reaction_flags(reactions_obj) -> List[str]:
    if not reactions_obj:
        return []
    s = str(reactions_obj)
    flags = []
    for r in REACTIONS:
        if r in s:
            flags.append(r)
    return flags

# ------------------------ CHANNEL <-> DB SYNC HELPERS ----------------

async def fetch_db_urls_ordered(cutoff: int) -> List[Tuple[int, str, int, Optional[int]]]:
    """
    Возвращает список кортежей (id, url, first_seen, sent_channel_msg_id) упорядоченных по first_seen asc
    Только записи с first_seen >= cutoff.
    """
    rows = await db_fetchall('SELECT id, url, first_seen, sent_channel_msg_id FROM urls WHERE first_seen >= ? ORDER BY first_seen ASC', (cutoff,))
    return rows

async def get_channel_url_list(limit=CHANNEL_FETCH_LIMIT) -> List[Tuple[int, str, int]]:
    """
    Возвращает список (msg.id, normalized_url, ts) из канала в порядке от старых к новым.
    """
    rows = await gather_channel_messages_with_urls(limit=limit)
    return [(mid, norm, ts) for (mid, text, norm, ts, reactions) in rows]

async def ensure_channel_matches_db(cutoff: int):
    """
    Убедиться, что содержимое канала соответствует DB (urls с first_seen >= cutoff) по набору и порядку.
    Алгоритм:
      1) получаем desired ordered list (по first_seen)
      2) получаем текущие сообщения в канале (ordered)
      3) если lists совпадают по последовательности urls -> просто обновляем sent_channel_msg_id в БД (если
         совпадает url -> записываем msg.id в sent_channel_msg_id)
      4) иначе -> очищаем канал и заново заливаем desired в правильном порядке (и записываем sent_channel_msg_id)
    """
    desired = await fetch_db_urls_ordered(cutoff)
    desired_urls = [row[1] for row in desired]
    channel_msgs = await get_channel_url_list()
    channel_urls = [norm for (_, norm, _) in channel_msgs]

    if channel_urls == desired_urls:
        logger.info("Channel already matches DB order — updating DB mapping of sent_channel_msg_id where needed.")
        # Обновим sent_channel_msg_id в БД, чтобы синхронизировать id'ы (если пользователь очистил частично и восстановил, или
        # если БД не знала текущий id)
        # Создадим mapping url -> msg.id
        url_to_msgid = {norm: mid for (mid, norm, ts) in channel_msgs}
        for url_id, url, first_seen, sent_msg in desired:
            mid = url_to_msgid.get(url)
            if mid and sent_msg != mid:
                try:
                    await db_execute('UPDATE urls SET sent_channel_msg_id=? WHERE id=?', (mid, url_id))
                except Exception:
                    logger.exception('Failed to update sent_channel_msg_id during sync')
        logger.info("DB mapped to current channel messages.")
        return

    # Если не совпадают — перезаписываем канал: очищаем и отправляем в правильном порядке
    logger.info("Channel content differs from DB desired list: rebuilding channel to match DB order.")
    # очистка
    try:
        try:
            target = int(TARGET_CHANNEL)
        except Exception:
            target = TARGET_CHANNEL
        msgs = await client.get_messages(entity=target, limit=CHANNEL_FETCH_LIMIT)
        ids = [m.id for m in msgs if getattr(m, 'id', None) is not None]
        BATCH = 100
        for i in range(0, len(ids), BATCH):
            batch = ids[i:i+BATCH]
            try:
                await client.delete_messages(entity=target, message_ids=batch)
                await asyncio.sleep(0.2)
            except Exception as e:
                logger.warning(f"Failed to delete batch while rebuilding channel: {e}")
        logger.info("Target channel cleared for rebuild.")
    except Exception:
        logger.exception("Failed to clear target channel during rebuild.")

    # Отправим в канале все desired в порядке и обновим sent_channel_msg_id
    sent = 0
    for url_id, url, first_seen, sent_msg in desired:
        try:
            new_mid = await send_channel_message(url, first_seen)
            if new_mid:
                try:
                    await db_execute('UPDATE urls SET sent_channel_msg_id=? WHERE id=?', (new_mid, url_id))
                except Exception:
                    logger.exception('Failed to update sent_channel_msg_id after sending during rebuild')
                sent += 1
            await asyncio.sleep(0.2)
        except Exception:
            logger.exception('Failed to send message during channel rebuild')
    logger.info(f"Rebuilt channel with {sent} messages to match DB.")

# ------------------------ STARTUP SEQUENCE ------------------------

async def init_db_tables():
    await db_execute('''
    CREATE TABLE IF NOT EXISTS urls (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        url TEXT UNIQUE,
        first_seen INTEGER,
        sent_channel_msg_id INTEGER,
        reacted_flags TEXT DEFAULT ''
    )
    ''')
    await db_execute('''
    CREATE TABLE IF NOT EXISTS occurrences (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        url_id INTEGER,
        peer TEXT,
        message_id INTEGER,
        message_date INTEGER,
        FOREIGN KEY(url_id) REFERENCES urls(id)
    )
    ''')

async def apply_reactions_from_channel_to_sources():
    logger.info('Starting reaction sync from channel...')
    rows = await gather_channel_messages_with_urls(limit=CHANNEL_FETCH_LIMIT)
    if not rows:
        logger.info('No messages found in channel for reaction sync.')
        return 0
    applied_total = 0
    for msg_id, text, norm, ts, reactions_obj in rows:
        flags = parse_reaction_flags(reactions_obj)
        if not flags:
            continue
        res = await db_fetchall('SELECT id, reacted_flags FROM urls WHERE url = ?', (norm,))
        if not res:
            continue
        url_id, reacted_flags_db = res[0]
        reacted_flags_db = reacted_flags_db or ''
        occs = await db_fetchall('SELECT peer, message_id FROM occurrences WHERE url_id = ?', (url_id,))
        for peer_str, occ_msg_id in occs:
            try:
                peer = int(peer_str)
            except Exception:
                peer = peer_str
            for r in flags:
                if r in reacted_flags_db:
                    continue
                ok = await try_send_reaction(peer, occ_msg_id, r)
                if ok:
                    applied_total += 1
                    await asyncio.sleep(0.25)
        if flags:
            new_flags = ''.join(sorted(set((reacted_flags_db or '') + ''.join(flags))))
            try:
                await db_execute('UPDATE urls SET reacted_flags=? WHERE id=?', (new_flags, url_id))
            except Exception:
                logger.exception('Failed to update reacted_flags in DB')
    logger.info(f'Reaction sync finished — applied approx: {applied_total} reactions')
    return applied_total

async def cleanup_old_links():
    logger.info('Starting cleanup of old links (>24h)...')
    now = int(time.time())
    cutoff = now - 24 * 3600
    rows = await db_fetchall('SELECT id, sent_channel_msg_id FROM urls WHERE first_seen < ?', (cutoff,))
    removed = 0
    for url_id, sent_msg_id in rows:
        if sent_msg_id:
            try:
                await delete_channel_message(sent_msg_id)
            except Exception:
                logger.exception('delete_channel_message failed during cleanup')
        try:
            await db_execute('DELETE FROM occurrences WHERE url_id=?', (url_id,))
            await db_execute('DELETE FROM urls WHERE id=?', (url_id,))
            removed += 1
            logger.info(f'Removed url id={url_id} older than 24h')
        except Exception:
            logger.exception('Failed to delete DB rows during cleanup')
    logger.info(f'Cleanup finished. Removed {removed} old urls.')
    return removed

async def scan_sources_and_insert_recent(cutoff: int):
    """
    Сканируем источники и вставляем occurrence'ы и новые urls для сообщений младше cutoff.
    НЕ отправляет в канал — только записывает в БД.
    Возвращает число новых url'ов.
    """
    logger.info('Scanning sources for recent links (to be inserted into DB)...')
    new_added = 0
    for src in SOURCES:
        try:
            msgs = await client.get_messages(entity=src, limit=MAX_MESSAGES_POLL)
        except Exception as e:
            logger.warning(f'Polling source {src} failed: {e}')
            continue
        msgs = list(reversed(msgs))
        for m in msgs:
            if not m or not getattr(m, 'message', None):
                continue
            try:
                msg_ts = int(m.date.replace(tzinfo=pytz.utc).timestamp())
            except Exception:
                continue
            if msg_ts < cutoff:
                continue
            text = m.message
            found = URL_RE.findall(text)
            if not found:
                continue
            for raw in found:
                url_norm = normalize_url(raw)
                res = await db_fetchall('SELECT id FROM urls WHERE url=?', (url_norm,))
                if res:
                    url_id = res[0][0]
                    try:
                        await db_execute('INSERT INTO occurrences(url_id, peer, message_id, message_date) VALUES (?,?,?,?)',
                                         (url_id, str(src), m.id, msg_ts))
                    except Exception:
                        # duplicate occurrences possible — игнорируем
                        pass
                else:
                    try:
                        lastrow = await db_execute('INSERT INTO urls(url, first_seen) VALUES (?,?)', (url_norm, msg_ts))
                        url_id = int(lastrow)
                        await db_execute('INSERT INTO occurrences(url_id, peer, message_id, message_date) VALUES (?,?,?,?)',
                                         (url_id, str(src), m.id, msg_ts))
                        new_added += 1
                    except Exception:
                        logger.exception('Failed to insert new url during scanning')
        await asyncio.sleep(0.25)
    logger.info(f'Scan finished. New urls inserted into DB: {new_added}')
    return new_added

async def initial_startup_sequence():
    """
    Порядок действий при старте:
      1) старт DB worker + создание таблиц
      2) синхронизация реакций (канал -> occurrences) — выполняется ДО удаления
      3) удаление старых записей из БД (>24ч)
      4) сверка/восстановление канала, чтобы порядок совпадал с БД (если флаг CLEAR_TARGET_CHANNEL_ON_START=True — очищаем и строим заново)
      5) скан источников и вставка новых ссылок в БД (только последние 24ч)
      6) после вставки — снова проверяем/восстанавливаем канал (чтобы учесть только что добавленные ссылки) и записываем sent_channel_msg_id
      7) rebuild Excel
    """
    asyncio.create_task(db_worker())
    await init_db_tables()

    now = int(time.time())
    cutoff = now - 24 * 3600

    # 1) reaction sync
    try:
        await apply_reactions_from_channel_to_sources()
    except Exception:
        logger.exception('Reaction sync failed during startup')

    # 2) cleanup old DB rows
    try:
        await cleanup_old_links()
    except Exception:
        logger.exception('Initial cleanup failed')

    # 3) if CLEAR flag True -> clear and then build using DB; else try to match DB and rebuild if mismatch
    try:
        if CLEAR_TARGET_CHANNEL_ON_START:
            # Очистим канал полностью (we use ensure_channel_matches_db which will clear+rebuild if mismatch)
            # But first ensure DB has current urls (maybe it's empty now) — we will insert recent from sources below,
            # so we want to postpone the rebuild until after scanning sources. For now, if DB non-empty, we will rebuild,
            # otherwise we just clear.
            try:
                try:
                    target = int(TARGET_CHANNEL)
                except Exception:
                    target = TARGET_CHANNEL
                msgs = await client.get_messages(entity=target, limit=CHANNEL_FETCH_LIMIT)
                ids = [m.id for m in msgs if getattr(m, 'id', None) is not None]
                BATCH = 100
                for i in range(0, len(ids), BATCH):
                    batch = ids[i:i+BATCH]
                    try:
                        await client.delete_messages(entity=target, message_ids=batch)
                        await asyncio.sleep(0.2)
                    except Exception as e:
                        logger.warning(f"Failed to delete batch while clearing channel at startup: {e}")
                logger.info("Target channel cleared (startup clear).")
            except Exception:
                logger.exception("Failed to clear target channel at startup.")
        else:
            # Если CLEAR=False — попытаемся убедиться, что содержимое совпадает, иначе перестроим
            await ensure_channel_matches_db(cutoff)
    except Exception:
        logger.exception('Channel clear/sync failed during startup')

    # 4) scan sources and insert recent into DB (do not send yet)
    try:
        await scan_sources_and_insert_recent(cutoff)
    except Exception:
        logger.exception('Initial source scan failed')

    # 5) теперь гарантированно синхронизируем канал с DB (после вставки новых ссылок)
    try:
        await ensure_channel_matches_db(cutoff)
    except Exception:
        logger.exception('Final ensure channel vs DB failed')

    # 6) rebuild excel
    try:
        await rebuild_excel_from_db(EXCEL_PATH)
    except Exception:
        logger.exception('Excel rebuild failed on startup')

    logger.info('Initial startup sequence finished')

# ------------------------ Event handler (new messages) ------------------------

@client.on(events.NewMessage(chats=SOURCES))
async def handler_new_message(event):
    msg = event.message
    text = msg.message or ''
    if not text:
        return
    found = URL_RE.findall(text)
    if not found:
        return
    msg_ts = int(msg.date.replace(tzinfo=pytz.utc).timestamp())
    for raw in found:
        url = normalize_url(raw)
        try:
            res = await db_fetchall('SELECT id FROM urls WHERE url=?', (url,))
            if res:
                url_id = res[0][0]
                try:
                    await db_execute('INSERT INTO occurrences(url_id, peer, message_id, message_date) VALUES (?,?,?,?)',
                                     (url_id, str(event.chat_id), msg.id, msg_ts))
                except Exception:
                    logger.debug('Insert occurrence failed or duplicate in handler_new_message')
            else:
                lastrow = await db_execute('INSERT INTO urls(url, first_seen) VALUES (?,?)', (url, msg_ts))
                url_id = int(lastrow)
                try:
                    await db_execute('INSERT INTO occurrences(url_id, peer, message_id, message_date) VALUES (?,?,?,?)',
                                     (url_id, str(event.chat_id), msg.id, msg_ts))
                except Exception:
                    logger.debug('Insert occurrence failed after url insert')
                # Отправляем новое сообщение (только если сообщение свежее 24ч)
                now = int(time.time())
                cutoff = now - 24 * 3600
                if msg_ts >= cutoff:
                    sent_id = None
                    try:
                        sent_id = await send_channel_message(url, msg_ts)
                    except Exception:
                        logger.exception('Failed to send to channel on new message')
                    if sent_id:
                        try:
                            await db_execute('UPDATE urls SET sent_channel_msg_id=? WHERE id=?', (sent_id, url_id))
                        except Exception:
                            logger.warning('Failed to update sent_channel_msg_id on new message')
        except Exception:
            logger.exception('Error processing new message entry')
    try:
        await rebuild_excel_from_db(EXCEL_PATH)
    except Exception:
        logger.exception('Excel rebuild failed after handling new message')

# ------------------------ MAIN / START ------------------------------

async def main():
    ensure_excel(EXCEL_PATH)
    await client.start(phone=PHONE)
    logger.info('Client started')
    await initial_startup_sequence()
    logger.info('Bot is running. Now listening for new messages...')
    try:
        await client.run_until_disconnected()
    finally:
        # корректно остановим db_worker
        await _db_queue.put(None)

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info('Interrupted by user, shutting down...')
    except Exception:
        logger.exception('Fatal error')
